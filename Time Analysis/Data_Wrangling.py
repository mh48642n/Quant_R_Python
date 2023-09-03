# -*- coding: utf-8 -*-
"""
Created on Mon Jun 12 22:32:19 2023

@author: marvi
"""
"""
Pulling data from columns (A:D)3:(A:D)6371 

"""

import openpyxl as op
import pandas as pd
import numpy as np


#pulls data and makes a specfic worksheet active, allowing us to to interact
bk = op.load_workbook("C:\\Users\\marvi\\OneDrive\\Documents\\GitHub\\Quant_R_Python\\Time Analysis\\Incubator_Project_-_Daily_Exchange_rates_1999_to_now.xlsx") 
bk.active = bk["Conversion Rates"]
sheet = bk.active

#iterates through column A from cell 12 to 6376 accounting for dates
dates = [sheet.cell(row=i,column=1).value for i in range(3,6371)]

#iterates through column B-D from cell 3 to 6371 accounting for rates 
DEXSEU_Euro = [sheet.cell(row=i,column=2).value for i in range(3,6371)]
DEXJPUS_Yen = [sheet.cell(row=i,column=3).value for i in range(3,6371)]
DEXCHUS_Yuan = [sheet.cell(row=i,column=4).value for i in range(3,6371)]
    

#creates dictionary out of the lists and then the dictionary is converted to df
dictt = {"dates": dates, "Euro": DEXSEU_Euro, "Yen": DEXJPUS_Yen, "Yuan": DEXCHUS_Yuan}
r_to_d = pd.DataFrame(dictt)

#Close this workbook
bk.close()

#Replaces all values called Not Available with an actual Nan value to make
#analysis in R a bit easier
r_to_d.replace(to_replace = ["Not Available"], value = np.nan, inplace = True)


r_to_d.to_excel('Rates_and_dates.xlsx', sheet_name = "Conversion Rates", index = False)


#adding the 10-year maturities minus 2-year maturities and 
#10 year maturity minus the 3 month maturities 
#then saving to Rates_and_dates.xlsx

def newsheet(data1, data2):
    wk1 = op.load_workbook(data1)
    wk2 = op.load_workbook(data2)
    sheet1 = wk1.active
    sheet2 = wk2.active
    
    dates = [sheet1.cell(row = i, column = 1).value for i in range(12, 6379)]
    yields_3m = [sheet1.cell(row = i, column = 2).value for i in range(12, 6379)] 
    yields_2y = [sheet2.cell(row = i, column = 2).value for i in range(12, 6379)] 
   
    
    dictt = {"Dates": dates, 
             "3mth_yields" : yields_3m, 
             "2yr_yields" : yields_2y}
    
    yields_to_d = pd.DataFrame(dictt)
    return yields_to_d

#Getting new dataframe from the newsheet function
ph1 = "C:\\Users\marvi\Downloads\T10Y3M.xlsx"
ph2 = "C:\\Users\marvi\Downloads\T10Y2Y.xlsx"
data = newsheet(ph1, ph2)

#Making sure NA values are still there
data.replace(to_replace = [""], value = np.nan, inplace = True)

#Saving dataframe, data, to a excel workbook
data.to_excel("Yields.xlsx", sheet_name = "Yield Rates", index = False)


del data, ph1, ph2
    

    
    
    
