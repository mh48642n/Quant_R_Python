# -*- coding: utf-8 -*-
"""
Created on Mon Jun 12 22:32:19 2023

@author: marvi
"""
"""
Pulling data from columns (A:D)3:(A:D)6371 

"""

import openpyxl as op
import pandas as pd
import numpy as np


#pulls data and makes a specfic worksheet active, allowing us to to interact
sheet = op.load_workbook("C:\\Users\\marvi\\OneDrive\\Documents\\GitHub\\Quant_R_Python\\Time Analysis\\Incubator_Project_-_Daily_Exchange_rates_1999_to_now.xlsx")
sheet.active = sheet["Conversion Rates"]
sheet = sheet.active


#iterates through column A from cell 12 to 6376 accounting for dates
dates = [sheet.cell(row=i,column=1).value for i in range(3,6371)]

#iterates through column B-D from cell 3 to 6371 accounting for rates 
DEXSEU_Euro = [sheet.cell(row=i,column=2).value for i in range(3,6371)]
DEXJPUS_Yen = [sheet.cell(row=i,column=3).value for i in range(3,6371)]
DEXCHUS_Yuan = [sheet.cell(row=i,column=4).value for i in range(3,6371)]
    

#creates dictionary out of the lists and then the dictionary is converted to df
dictt = {"dates": dates, "Euro": DEXSEU_Euro, "Yen": DEXJPUS_Yen, "Yuan": DEXCHUS_Yuan}
r_to_d = pd.DataFrame(dictt)

#Replaces all values called Not Available with an actual Nan value to make
#analysis in R a bit easier
r_to_d.replace(to_replace = ["Not Available"], value = np.nan, inplace = True)


r_to_d.to_excel('Rates_and_dates.xlsx', sheet_name = "Conversion Rates", index = False)




